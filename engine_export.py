"""
engine_export.py — Premium Dynamic Exporter (v4.0)
Constructs formatted multi-sheet Excel workbooks matching dashboard totals exactly.
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Premium Corporate Palette
NAVY   = "1F3864"; BLUE   = "2E75B6"; LTBLUE = "DBEAFE"
RED    = "C00000"; LTRED  = "FEE2E2"; AMBER  = "C55A11"
LTAMB  = "FEF3C7"; GREEN  = "166534"; LTGRN  = "DCFCE7"
GRAY1  = "F8FAFC"; GRAY2  = "E2E8F0"; GRAY3  = "64748B"
WHITE  = "FFFFFF"

# Defect subcategories
HIGH_SUBCATS = ["Defective Product", "Damaged Product", "Low Quality Product", "Order Delay", "Order Not Shipped"]


def _s(style="thin", color="CBD5E0"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _f(c): 
    return PatternFill("solid", start_color=c, end_color=c)


def _c(ws, r, c, v, bold=False, align="left", nf=None, fg="1E293B", bg=None, sz=9, wrap=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Calibri", bold=bold, size=sz, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _s()
    if nf:  
        cell.number_format = nf
    if bg:  
        cell.fill = _f(bg)
    return cell


def _h(ws, r, c, v, bg=NAVY, fg=WHITE, sz=10, align="center", wrap=True, height=None):
    cell = _c(ws, r, c, v, bold=True, align=align, fg=fg, bg=bg, sz=sz, wrap=wrap)
    if height: 
        ws.row_dimensions[r].height = height
    return cell


def _title(ws, title, sub, cols):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    c = ws["A1"]
    c.value = title
    c.font  = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill  = _f(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{get_column_letter(cols)}2")
    c2 = ws["A2"]
    c2.value = sub
    c2.font  = Font(name="Calibri", size=9, italic=True, color=GRAY3)
    c2.fill  = _f(GRAY1)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16


def _autofit(ws):
    """Enables gridlines, freeze panes on A5, and auto-fits columns dynamically."""
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A5"
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)


def _impact_style(impact):
    return {"CRITICAL": (LTRED, RED), "HIGH": (LTAMB, AMBER),
            "MEDIUM": (LTBLUE, BLUE), "LOW": (LTGRN, GREEN)}.get(str(impact).upper(), (WHITE, "000000"))


def dump_sheet(wb, sheet_name, title_txt, period, headers, df_data, formats=None):
    """Generates standard tabular layouts with dynamic format masks."""
    ws = wb.create_sheet(sheet_name)
    _title(ws, title_txt, f"Filter Context: {period}", len(headers))
    for ci, h in enumerate(headers, 1):
        _h(ws, 4, ci, h)
        
    idx = 5
    if not df_data.empty:
        for _, row in df_data.iterrows():
            ws.row_dimensions[idx].height = 16
            for ci, col in enumerate(df_data.columns, 1):
                val = row[col]
                nf = formats.get(col) if formats else None
                _c(ws, idx, ci, val, align="center" if ci > 1 else "left", nf=nf)
            idx += 1
    _autofit(ws)
    return ws


def _sheet_exec(wb, kpis, brand_sum, subcat_sum, period, orig, final, val_ok):
    """Generates styled executive scorecards and tables."""
    ws = wb.create_sheet("Executive Summary")
    _title(ws, "EXECUTIVE SUMMARY — PERFORMANCE OVERVIEW", f"Period: {period}", 6)
    
    kpis_list = [
        ("Base Volume", f"{kpis['total_del']:,}", NAVY),
        ("Tickets Base", f"{kpis['total_tick']:,}", RED),
        ("Escalation Rate %", f"{kpis['overall_esc']}%", AMBER),
        ("Defect Rate %", f"{kpis['overall_defect']}%", GREEN),
        ("Peak Week", kpis.get('spike_week', '—'), BLUE)
    ]
    for i, (lbl, val, bg) in enumerate(kpis_list, 1):
        _c(ws, 4, i, lbl, bold=True, bg="F1F5F9", fg=NAVY, sz=8, align="center")
        _c(ws, 5, i, val, bold=True, fg="1E293B", sz=14, align="center")
    
    _c(ws, 4, 6, "Validation Checks", bold=True, bg="F1F5F9", fg=NAVY, sz=8, align="center")
    _c(ws, 5, 6, "PASS ✅" if val_ok else "FAIL ❌", bold=True, fg=GREEN if val_ok else RED, sz=12, align="center")
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 26

    # Brand Escalation Table
    ws.cell(row=7, column=1, value="TOP BRAND PROFILES SUMMARY").font = Font(name="Calibri", bold=True, size=11, color=NAVY)
    _h(ws, 8, 1, "Brand"); _h(ws, 8, 2, "Orders"); _h(ws, 8, 3, "Tickets"); _h(ws, 8, 4, "Escalation Rate %"); _h(ws, 8, 5, "Defect Rate %"); _h(ws, 8, 6, "Impact Status")
    
    r = 9
    for _, row in brand_sum.head(10).iterrows():
        _c(ws, r, 1, row["brand"], bold=True)
        _c(ws, r, 2, row["delivered"], align="center", nf="#,##0")
        _c(ws, r, 3, row["tickets"], align="center", nf="#,##0")
        
        ibg, ifg = _impact_style(row["impact"])
        _c(ws, r, 4, row["esc_pct"]/100, align="center", nf="0.0%", bg=ibg, fg=ifg, bold=True)
        _c(ws, r, 5, row["defect_rate"]/100, align="center", nf="0.0%")
        _c(ws, r, 6, row["impact"], align="center", bg=ibg, fg=ifg, bold=True)
        ws.row_dimensions[r].height = 16
        r += 1

    # Issue Table
    ws.cell(row=r+1, column=1, value="PRIMARY ISSUE DRIVERS ANALYSIS").font = Font(name="Calibri", bold=True, size=11, color=NAVY)
    r += 2
    _h(ws, r, 1, "Issue Category"); _h(ws, r, 2, "Tickets"); _h(ws, r, 3, "Share %")
    
    for _, row in subcat_sum.head(5).iterrows():
        r += 1
        _c(ws, r, 1, row["subcat_final"], bold=True)
        _c(ws, r, 2, row["count"], align="center", nf="#,##0")
        _c(ws, r, 3, row["pct"]/100, align="center", nf="0.0%")
        ws.row_dimensions[r].height = 16
        
    _autofit(ws)


def _sheet_brand_analytics_generic(wb, name, title, brand_sum, period):
    dump_sheet(wb, name, title, period,
               ["Brand", "Orders Volume", "Volume % Share", "Tickets", "Ticket % Share", "Escalation %", "Defect %", "Weighted Escalation %", "Confidence %", "Primary Issue", "Impact"],
               brand_sum[["brand", "delivered", "del_share", "tickets", "tick_share", "esc_pct", "defect_rate", "weighted_esc", "confidence", "Top Escalation Driver", "impact"]],
               {"delivered": "#,##0", "tickets": "#,##0", "esc_pct": "0.0%", "defect_rate": "0.0%", "weighted_esc": "0.0%", "confidence": '0"%"'})


def _sheet_product_analytics_generic(wb, name, title, prod_sum, period):
    dump_sheet(wb, name, title, period,
               ["Brand", "Product Name", "Orders Volume", "Tickets", "Escalation %", "Weighted Escalation %", "Confidence %", "Primary Source Month", "Same Month", "Prev Month", "Older Tickets", "Aging Category", "Impact"],
               prod_sum[["brand", "canonical_product", "delivered", "tickets", "esc_pct", "weighted_esc", "confidence", "Primary Ticket Source Month", "Same Month Tickets", "Previous Month Tickets", "Older Tickets", "Ticket Aging Category", "impact"]],
               {"delivered": "#,##0", "tickets": "#,##0", "esc_pct": "0.0%", "weighted_esc": "0.0%", "confidence": '0"%"'})


def _sheet_issue_generic(wb, name, title, subcat_sum, period):
    dump_sheet(wb, name, title, period,
               ["Issue Category", "Tickets", "Share %", "Severity Tier"],
               subcat_sum[["subcat_final", "count", "pct", "tier"]] if not subcat_sum.empty else pd.DataFrame(columns=["subcat_final", "count", "pct", "tier"]),
               {"count": "#,##0", "pct": "0.0%"})


def _sheet_validation_report(wb, orig, final, val_ok, n_unmapped, n_need_details, n_not_found, period):
    ws = wb.create_sheet("Validation Report")
    _title(ws, "DATASET INTEGRITY VALIDATION PANEL", f"Period: {period}", 5)
    
    stats = [
        ("Original Tickets Uploaded", orig, "blue"),
        ("Final Reconciled Tickets", final, "green" if val_ok else "red"),
        ("Unmapped Brand Tickets Apportioned", n_unmapped, "purple"),
        ("Need Details Tickets Restructured", n_need_details, "purple"),
        ("Not Found Tickets Restructured", n_not_found, "purple"),
        ("Data Sync Status", "PASS ✅" if val_ok else "FAIL ❌", "green" if val_ok else "red")
    ]
    
    r = 4
    for label, val, color_name in stats:
        ws.merge_cells(f"A{r}:C{r}")
        ws.merge_cells(f"D{r}:E{r}")
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=4, value=val)
        
        col = RED if "FAIL" in str(val) or color_name == "red" else GREEN if "PASS" in str(val) or color_name == "green" else NAVY
        lc.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
        vc.font = Font(name="Calibri", bold=True, size=11, color=col)
        
        ws.row_dimensions[r].height = 16
        r += 1
        
    _autofit(ws)


def generate_excel_report(kpis, brand_sum, prod_sum, subcat_sum,
                          weekly_trends, redist_summary, cohort_report,
                          comp_df_brand, comp_df_prod, registry, tick_df, del_df,
                          orig_tickets, final_tickets, val_ok, period="All Data"):
    """
    Constructs a dynamic workbook aligning values cleanly with active dataset metrics.
    Employs unique Order ID denominators across sheets for mathematical correctness.
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default active sheet

    subcat_col = "subcat_final" if "subcat_final" in tick_df.columns else "raw_subcat"
    order_col = "order_id" if "order_id" in del_df.columns else "zop_id"

    # Dynamic pre-calculation segments for Pre, Post, and Combined Summaries
    post_del = del_df[del_df["is_delivered"] == True]
    post_tick = tick_df[tick_df["ticket_category"] == "POST_DELIVERY"]
    
    pre_del = del_df.copy()
    pre_tick = tick_df[tick_df["ticket_category"] == "PRE_DELIVERY"]
    
    from engine_analytics import compute_brand_summary, compute_product_summary, compute_subcat_summary
    
    pre_brand = compute_brand_summary(del_df=pre_del, tick_df=pre_tick) if not pre_tick.empty else brand_sum.copy()
    post_brand = compute_brand_summary(del_df=post_del, tick_df=post_tick) if not post_tick.empty else brand_sum.copy()
    
    pre_prod = compute_product_summary(del_df=pre_del, tick_df=pre_tick) if not pre_tick.empty else prod_sum.copy()
    post_prod = compute_product_summary(del_df=post_del, tick_df=post_tick) if not post_tick.empty else prod_sum.copy()
    
    pre_subcat = compute_subcat_summary(pre_tick)
    post_subcat = compute_subcat_summary(post_tick)

    # Compute exact unique orders denominators for segment scorecards
    pre_orders_count = pre_del[order_col].nunique() if not pre_del.empty else 0
    pre_tickets_count = len(pre_tick)
    pre_esc_rate = round((pre_tickets_count / max(pre_orders_count, 1)) * 100, 2)
    pre_defect_count = len(pre_tick[pre_tick[subcat_col].isin(HIGH_SUBCATS)]) if not pre_tick.empty else 0
    pre_defect_rate = round((pre_defect_count / max(pre_orders_count, 1)) * 100, 2)
    
    pre_kpis = {
        "total_del": pre_orders_count,
        "total_tick": pre_tickets_count,
        "overall_esc": pre_esc_rate,
        "overall_defect": pre_defect_rate,
        "spike_week": kpis.get("spike_week", "—")
    }

    post_orders_count = post_del[order_col].nunique() if not post_del.empty else 0
    post_tickets_count = len(post_tick)
    post_esc_rate = round((post_tickets_count / max(post_orders_count, 1)) * 100, 2)
    post_defect_count = len(post_tick[post_tick[subcat_col].isin(HIGH_SUBCATS)]) if not post_tick.empty else 0
    post_defect_rate = round((post_defect_count / max(post_orders_count, 1)) * 100, 2)
    
    post_kpis = {
        "total_del": post_orders_count,
        "total_tick": post_tickets_count,
        "overall_esc": post_esc_rate,
        "overall_defect": post_defect_rate,
        "spike_week": kpis.get("spike_week", "—")
    }

    comb_orders_count = del_df[order_col].nunique() if not del_df.empty else 0
    comb_tickets_count = len(tick_df)
    comb_esc_rate = round((comb_tickets_count / max(comb_orders_count, 1)) * 100, 2)
    comb_defect_count = len(tick_df[tick_df[subcat_col].isin(HIGH_SUBCATS)]) if not tick_df.empty else 0
    comb_defect_rate = round((comb_defect_count / max(comb_orders_count, 1)) * 100, 2)
    
    comb_kpis = {
        "total_del": comb_orders_count,
        "total_tick": comb_tickets_count,
        "overall_esc": comb_esc_rate,
        "overall_defect": comb_defect_rate,
        "spike_week": kpis.get("spike_week", "—")
    }

    # ── SHEET 1: Executive Summary ──
    _sheet_exec(wb, kpis, brand_sum, subcat_sum, period, orig_tickets, final_tickets, val_ok)
    
    # ── SHEET 2: Pre Delivery Summary ──
    _sheet_exec(wb, pre_kpis, pre_brand, pre_subcat, period, orig_tickets, len(pre_tick), val_ok)
    wb.worksheets[-1].title = "Pre Delivery Summary"
    
    # ── SHEET 3: Post Delivery Summary ──
    _sheet_exec(wb, post_kpis, post_brand, post_subcat, period, orig_tickets, len(post_tick), val_ok)
    wb.worksheets[-1].title = "Post Delivery Summary"
    
    # ── SHEET 4: Combined Summary ──
    _sheet_exec(wb, comb_kpis, brand_sum, subcat_sum, period, orig_tickets, len(tick_df), val_ok)
    wb.worksheets[-1].title = "Combined Summary"
    
    # ── SHEETS 5 - 7: Brand Analytics ──
    _sheet_brand_analytics_generic(wb, "Pre Brand Analytics", "PRE-DELIVERY BRAND PERFORMANCE SUMMARY", pre_brand, period)
    _sheet_brand_analytics_generic(wb, "Post Brand Analytics", "POST-DELIVERY BRAND PERFORMANCE SUMMARY", post_brand, period)
    _sheet_brand_analytics_generic(wb, "Combined Brand Analytics", "COMBINED BRAND PERFORMANCE SUMMARY", brand_sum, period)
    
    # ── SHEETS 8 - 10: Product Analytics ──
    _sheet_product_analytics_generic(wb, "Pre Product Analytics", "PRE-DELIVERY PRODUCT SUMMARY", pre_prod, period)
    _sheet_product_analytics_generic(wb, "Post Product Analytics", "POST-DELIVERY PRODUCT SUMMARY", post_prod, period)
    _sheet_product_analytics_generic(wb, "Combined Product Analytics", "COMBINED PRODUCT SUMMARY", prod_sum, period)
    
    # ── SHEETS 11 - 13: Issue Breakdowns ──
    _sheet_issue_generic(wb, "Pre Issue Breakdown", "PRE-DELIVERY ISSUE SOURCE CLASSIFICATION", pre_subcat, period)
    _sheet_issue_generic(wb, "Post Issue Breakdown", "POST-DELIVERY ISSUE SOURCE CLASSIFICATION", post_subcat, period)
    _sheet_issue_generic(wb, "Combined Issue Breakdown", "COMBINED ISSUE SOURCE CLASSIFICATION", subcat_sum, period)
    
    # ── SHEET 14: Weekly Trends ──
    ws = wb.create_sheet("Weekly Trends")
    _title(ws, "WEEKLY TREND ANALYSIS OVERVIEW", f"Period: {period}", 7)
    _h(ws, 4, 1, "Week"); _h(ws, 4, 2, "Orders"); _h(ws, 4, 3, "Tickets"); _h(ws, 4, 4, "Escalation %"); _h(ws, 4, 5, "WoW Tickets"); _h(ws, 4, 6, "WoW Escalation %"); _h(ws, 4, 7, "Alert Status")
    r = 5
    for _, row in weekly_trends.iterrows():
        _c(ws, r, 1, row["Week"], bold=True)
        _c(ws, r, 2, row["Delivered"], align="center", nf="#,##0")
        _c(ws, r, 3, row["Tickets"], align="center", nf="#,##0")
        _c(ws, r, 4, row["Esc %"]/100, align="center", nf="0.0%", bold=True)
        _c(ws, r, 5, row["WoW Change Tickets"], align="center", nf="+0;-0;0")
        _c(ws, r, 6, row["WoW Change Esc %"]/100, align="center", nf="+0.0%;-0.0%;0%")
        _c(ws, r, 7, row["Spike Alert"], align="center")
        r += 1
    _autofit(ws)

    # ── SHEET 15: Validation Report ──
    n_unmapped_calc = int(tick_df["_redistributed"].sum()) if "_redistributed" in tick_df.columns else 0
    n_not_found_calc = int((tick_df["raw_subcat"] == "Not Found").sum()) if "raw_subcat" in tick_df.columns else 0
    n_need_details_calc = int((tick_df["raw_subcat"] == "Need Details").sum()) if "raw_subcat" in tick_df.columns else 0

    _sheet_validation_report(
        wb, orig_tickets, final_tickets, val_ok, 
        n_unmapped=n_unmapped_calc, 
        n_need_details=n_need_details_calc, 
        n_not_found=n_not_found_calc, 
        period=period
    )

    # ── SHEET 16: Redistribution Summary ──
    dump_sheet(wb, "Redistribution Summary", "TICKET APPORTIONMENT LOGIC REDISTRIBUTION AUDIT", period,
               ["Brand", "Allocation Weight", "Brand NF Absorbed", "Subcat NF Absorbed", "Need Details Absorbed"],
               redist_summary, {})

    # ── SHEET 17: Product Registry ──
    reg_df = registry.summary_df()
    dump_sheet(wb, "Product Registry", "CANONICAL PRODUCT REGISTRY — MAPPED GROUPS", period,
               ["Brand", "Canonical Product", "SKU", "Merged Variants", "Delivered Orders", "Tickets"],
               reg_df, {"Delivered Orders": "#,##0", "Tickets": "#,##0"})

    # ── SHEET 18: Operational Action Summary ──
    ws_act = wb.create_sheet("Operational Action Summary")
    _title(ws_act, "OPERATIONAL ACTIONS & STRATEGIC RECOMMENDATIONS", f"Period: {period}", 3)
    _h(ws_act, 4, 1, "Risk Level"); _h(ws_act, 4, 2, "Observed Trend Scenario"); _h(ws_act, 4, 3, "Mandatory SLA Action Plan")
    recs = [
        ("CRITICAL", "Brand escalation exceeds 7% with high defect volume", "🚨 HALT support log Dispatch — trigger vendor-level batches inspections"),
        ("HIGH RISK", "WoW escalation spike detected in current week trend", "⚠️ Reduce distribution flows — initiate programmatic root cause checks"),
        ("MEDIUM", "Emerging same-month risk profiles rising", "👁 Establish SLA metrics checks — implement active cohort checks")
    ]
    for idx, (risk, trend, act) in enumerate(recs, 5):
        _c(ws_act, idx, 1, risk, bold=True, bg=LTRED if "CRITICAL" in risk else LTAMB, fg=RED if "CRITICAL" in risk else AMBER)
        _c(ws_act, idx, 2, trend)
        _c(ws_act, idx, 3, act)
    _autofit(ws_act)

    # ── TIME COMPARISON SHEETS ──
    dump_sheet(wb, "Month Comparison", "MONTHLY ESCALATION COMPARISON REPORT", period,
               ["Delivery Month", "Orders", "Tickets", "Escalation %"],
               cohort_report[["Delivery Month", "delivered", "tickets", "esc_pct"]], {"delivered": "#,##0", "tickets": "#,##0", "esc_pct": "0.0%"})

    dump_sheet(wb, "Brand Comparison", "BRAND HISTORICAL COMPARE (MoM)", period,
               ["Brand", "Month A Esc %", "Month B Esc %", "Difference", "Escalation Trend"],
               comp_df_brand, {"Month A Esc %": "0.0%", "Month B Esc %": "0.0%"})

    dump_sheet(wb, "Product Comparison", "PRODUCT HISTORICAL COMPARE (MoM)", period,
               ["Brand", "Product", "Month A Esc %", "Month B Esc %", "Difference", "Escalation Trend"],
               comp_df_prod, {"Month A Esc %": "0.0%", "Month B Esc %": "0.0%"})

    dump_sheet(wb, "Ticket Attribution Analysis", "TICKET ATTRIBUTION ANALYSIS REPORT", period,
               ["Brand", "Product", "Primary Source Month", "Same Month Tickets", "Previous Month Tickets", "Older Tickets"],
               prod_sum[["brand", "canonical_product", "Primary Ticket Source Month", "Same Month Tickets", "Previous Month Tickets", "Older Tickets"]], {})

    dump_sheet(wb, "Ticket Aging Analysis", "TICKET AGING ANALYSIS MATRIX", period,
               ["Brand", "Product", "Same Month", "Prev Month", "Older Tickets", "Aging Category"],
               prod_sum[["brand", "canonical_product", "Same Month Tickets", "Previous Month Tickets", "Older Tickets", "Ticket Aging Category"]], {})

    dump_sheet(wb, "Delivery Cohorts", "DELIVERY COHORT ANALYSIS TRENDS", period,
               ["Delivery Month", "Orders", "Tickets", "Cohort Escalation Rate"],
               cohort_report[["Delivery Month", "delivered", "tickets", "esc_pct"]], {"delivered": "#,##0", "tickets": "#,##0", "esc_pct": "0.0%"})

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
